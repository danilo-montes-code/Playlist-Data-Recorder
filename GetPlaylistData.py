import openpyxl
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
import win32com.client


'''
Converts a passed in rgb tuple to hexadecimal, useful for changing the background of the top row on the excel sheet, making the column headers more distinguishable from other data

Args:
    rgb: a tuple a rgb values corresponding to a color

Return:
    the hexadecmial value of the passed in rgb value
'''
def rgb_to_hex(rgb):
    bgr = (rgb[2], rgb[1], rgb[0])
    str_value = '%02x%02x%02x' % bgr
    return int(str_value, 16)


'''
Checks a given playlist for a track.

Args:
    track: id of the track to search for
    playlist: playlist dictionary that will be searched for the track
    row: row that the song is in on the excel sheet
    col: column that the playlist is in on the excel sheet
'''
def check_playlist_for_track(track, playlist, row, col):
    for playlist_song in playlist['songs']:
        if playlist_song['track']['id'] == track['id']:
            print(f'Found song in {playlist["name"]}')
            sheet.cell(row, col).value = 'X'
            break


'''
Loops to write the data on the excel sheet
'''
def record_data_on_sheet():
    for row, og_track_data in enumerate(main_playlist['songs']):  # for every track
        print()
        print(f'Searching for {og_track_data["track"]["name"]}...')
        for col, playlist in enumerate(sub_playlists):  # for every other playlist
            check_playlist_for_track(og_track_data['track'], playlist, index+row+2, col+3)


'''
Gets the data and writes it to the excel sheet
'''
def write_data():
    global index, main_playlist, sheet
    book = openpyxl.load_workbook(sheet_file_location)
    sheet = book.active
    for i in range(0, main_playlist['number_of_songs'], 100):
        index = i
        shift_main_playlist_songs()
        write_to_first_two_columns()
        record_data_on_sheet()
        main_playlist['songs'] = []
    book.save(sheet_file_location)


'''
Writes the songs and their artists in the first and second columns respectively
'''
def write_to_first_two_columns():
    global sheet
    # puts the songs and their artists in the first and second columns respectively
    temp_artists = ''
    multiple_artists = False
    print(len(main_playlist['songs']))
    for i, song in enumerate(main_playlist['songs']):
        sheet.cell(index + i + 2, 1).value = song['track']['name']
        for artist in song['track']['artists']:
            if multiple_artists:
                temp_artists += ', ' + artist['name']
            else:
                temp_artists += artist['name']
            multiple_artists = True
        sheet.cell(index + i + 2, 2).value = temp_artists
        temp_artists = ''
        multiple_artists = False


'''
Appends to the list containing the dictionaries of all the playlists besides the main one

Args:
    playlist: the playlist to be appended to the list

Return:
    playlist dictionary after adding its tracks to said dictionary
'''
def create_sub_playlist_list(playlist):
    playlist_dict = {'name': playlist['name'],
                     'id': playlist['id'],
                     'number_of_songs': playlist['tracks']['total'],
                     'songs': []}

    results = sp.user_playlist(username, playlist['id'], fields="tracks,next")
    tracks = results['tracks']
    return add_tracks_to_list(tracks, playlist_dict)


'''
Adds tracks to the dictionary of the playlist that is passed in

Args:
    tracks: the tracks to be added to the dictionary
    playlist_dict: the playlist dictionary to add the songs to

Return:
    playlist_dict: the created playlist dictionary
'''
def add_tracks_to_list(tracks, playlist_dict):
    for track in tracks['items']:
        playlist_dict['songs'].append(track)
    while tracks['next']:
        tracks = sp.next(tracks)
        add_tracks_to_list(tracks, playlist_dict)
    return playlist_dict


'''
Shifts the songs down in the main playlist object and adds the songs to the main playlist dictionary
'''
def shift_main_playlist_songs():
    global main_playlist
    main_playlist_object = sp.user_playlist(username, main_playlist['id'], fields="tracks")
    playlist_tracks = main_playlist_object['tracks']
    shift = int(index/100)
    for i in range(0, shift):
        playlist_tracks = sp.next(playlist_tracks)
    for track in playlist_tracks['items']:
        main_playlist['songs'].append(track)


'''
Makes the sheet

Args:
    sheet_name: the user desired name of the excel sheet
'''
def make_sheet(sheet_name):
    global sheet_file_location, sheet

    # uses win32 to create the sheet and get the file path
    excel = win32com.client.Dispatch("Excel.Application")
    wb = excel.Workbooks.Add()
    ws = excel.ActiveSheet
    ws.SaveAs(sheet_name + '.xlsx')
    sheet_file_location = wb.Path + '\\' + sheet_name + '.xlsx'
    excel.Application.Quit()

    # uses openpyxl to set up the header row
    book = openpyxl.load_workbook(sheet_file_location)
    sheet = book.active

    # makes the header row
    sheet.cell(1, 1).value = main_playlist['name']
    sheet.cell(1, 2).value = "Artist(s)"
    for i, playlist in enumerate(sub_playlists):
        sheet.cell(1, i + 3).value = playlist['name']

    # saves the excel file's contents
    book.save(sheet_file_location)


'''
Formats the excel sheet
'''
def format_cells():
    # opens sheet with win32 for formatting
    excel = win32com.client.Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(sheet_file_location)
    ws = excel.ActiveSheet
    ws.Columns.Borders(11).LineStyle = 1

    # Auto re-sizes the columns, centers the text, and bolds and yellows the top row
    ws.Columns.AutoFit()
    ws.Columns.Style.HorizontalAlignment = -4108
    ws.Rows(1).Font.Bold = True
    ws.Rows(1).Interior.color = rgb_to_hex((255, 255, 0))

    # sets the borders to make the sheet easier to read
    ws.Rows(1).Borders.LineStyle = 1
    ws.Columns.Borders(11).LineStyle = 1
    # ws.Range(ws.Cells(2, 1), ws.Cells(last_row, last_col)).Sort(Key1=ws.Range(ws.Cells(2, 1),
    # ws.Cells(last_row, 1)), Order1=1, Orientation=2)

    # saves and quits
    wb.Save()
    excel.Application.Quit()


'''
Sets the global variables for use in the excel sheet

Args:
    main_playlist_name: name of the playlist
    all_playlists: list of all the playlists (seperated into 50-sized lists)

'''
def set_script_variables(main_playlist_name, all_playlists):
    global main_playlist, sub_playlists
    main_playlist_object = None
    # gets the main playlist that has all the songs and adds the songs into a dictionary
    for playlist_set in all_playlists:
        for playlist in playlist_set['items']:
            if playlist['owner']['id'] == username and playlist['name'] == main_playlist_name:
                main_playlist_object = playlist
                break
    try:
        main_playlist = {'name': main_playlist_object['name'],
                         'id': main_playlist_object['id'],
                         'number_of_songs': main_playlist_object['tracks']['total'],
                         'songs': []}
    except TypeError:
        print(f'Error: "{main_playlist_name}" is not amongst your public playlists.')
        main()

    # gets every other playlist and makes dictionaries for them
    sub_playlist_objects = []
    for playlist_set in all_playlists:
        for playlist in playlist_set['items']:
            if playlist['owner']['id'] == username and playlist['id'] != main_playlist['id']:
                sub_playlist_objects.append(playlist)
    for playlist in sub_playlist_objects:
        sub_playlists.append(create_sub_playlist_list(playlist))


def main():
    nan = True
    number_of_playlists = 0
    main_playlist_name = input('What is the name of your main playlist? (case sensitive)\n')
    while nan:
        try:
            number_of_playlists = int(input('How many public playlists do you have?\n'))
        except ValueError:
            print('Error: Please enter an integer number of public playlists')
        else:
            nan = False
    sheet_name = input('What do you want to name your excel sheet?\n')
    # puts all the playlists of the user into a list of lists (playlists from method call are
    # in sets of up to 50-sized lists, so this list is a list of those lists)
    all_playlists = []
    for i in range(0, number_of_playlists, 50):
        all_playlists.append(sp.user_playlists(username, offset=i))

    # sets up the global vars for excel sheet writing
    set_script_variables(main_playlist_name, all_playlists)

    # sets up the sheet object for writing
    make_sheet(sheet_name)  # TODO ask what color to make the top row

    # writes the data to the excel sheet and formats the cells
    write_data()
    format_cells()


if __name__ == '__main__':
    # sets up the credentials for the spotipy object
    client_credentials_manager = SpotifyClientCredentials(client_id='id',
                                                          client_secret='secret')
    sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)

    # sets up the variables that will be used throughout the script
    main_playlist = {}  # dictionary of the main playlist
    sub_playlists = []  # list of dictionaries of the sub playlists
    sheet_file_location = ''  # location of the excel file
    sheet = None  # excel sheet object
    index = 0  # buffer index to put shift down where the data is written on the sheet

    # asks the user for their username (hopefully temporary), taken from spotify uri
    user = input('What is your spotify uri? '
                 '(right click on your profile name when viewing your profile, hover over share, '
                 'and click the last option)\n')
    username = user[13:]
    main()
    # 1WKZ1xpg8BnmmPgTDDCrI4 - Monarchy
