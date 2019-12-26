import openpyxl  # does the bulk of the writing
import win32com.client  # does the formatting

excel_sheet = None
excel_sheet_file_location = ''

'''
Creates the sheet using win32com. Then sets up the header row.

Args:
    sheet_name: the user desired name of the excel sheet
'''
def create_sheet(name, top_row_color):
    global excel_sheet_file_location, excel_sheet

    # uses win32 to create the sheet and get the file path
    excel = win32com.client.Dispatch("Excel.Application")
    wb = excel.Workbooks.Add()
    ws = excel.ActiveSheet
    ws.SaveAs(sheet_name + '.xlsx')
    excel_sheet_file_location = wb.Path + '\\' + name + '.xlsx'
    excel.Application.Quit()

    # uses openpyxl to set up the header row
    book = openpyxl.load_workbook(excel_sheet_file_location)
    excel_sheet = book.active

    # makes the header row
    excel_sheet.cell(1, 1).value = main_playlist['name']
    excel_sheet.cell(1, 2).value = "Artist(s)"
    for i, playlist in enumerate(sub_playlists):
        excel_sheet.cell(1, i + 3).value = playlist['name']

    # saves the excel file's contents
    book.save(exce_sheet_file_location)


'''
Writes passed in data to a cell

Args:
    text: What is to be put in the cell
    row: The row number of the desired cell
    col: The column number of the desired cell
'''
def write_to_cell(text, row, col):
    pass


'''
Formats the excel sheet
'''
def format_cells(top_row_color):
    rgb_top_row_color = name_to_rgb(top_row_color)

    # opens sheet with win32 for formatting
    excel = win32com.client.Dispatch("Excel.Application")
    wb = excel.Workbooks.Open(sheet_file_location)
    ws = excel.ActiveSheet
    ws.Columns.Borders(11).LineStyle = 1

    # Auto re-sizes the columns, centers the text, and bolds and yellows the top row
    ws.Columns.AutoFit()
    ws.Columns.Style.HorizontalAlignment = -4108
    ws.Rows(1).Font.Bold = True
    ws.Rows(1).Interior.color = rgb_to_hex(rgb_top_row_color)

    # sets the borders to make the sheet easier to read
    ws.Rows(1).Borders.LineStyle = 1
    ws.Columns.Borders(11).LineStyle = 1
    # ws.Range(ws.Cells(2, 1), ws.Cells(last_row, last_col)).Sort(Key1=ws.Range(ws.Cells(2, 1),
    # ws.Cells(last_row, 1)), Order1=1, Orientation=2)

    # saves and quits
    wb.Save()
    excel.Application.Quit()

'''
Converts a color name to its rgb value

Args:
    color_name: the name of the color

Return:
    A list of the rgb values of the color
'''
def name_to_rgb(color_name):
    if color_name.lower() == 'black':
        return [255,255,255]
    elif color_name.lower() == 'yellow':
        return [255,255,0]



'''
Converts a passed in rgb tuple to hexadecimal, useful for changing the background of the top row on the excel sheet, making the column headers more distinguishable from other data

Args:
    rgb: a tuple a rgb values corresponding to a color

Return:
    the hexadecmial value of the passed in rgb value
'''
def rgb_to_hex(rgb):
    og = (rgb[2], rgb[1], rgb[0])
    str_value = '%02x%02x%02x' % og
    return int(str_value, 16)
